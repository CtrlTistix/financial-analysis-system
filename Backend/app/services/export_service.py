import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from io import BytesIO
from typing import Dict, List

class ExportService:
    def __init__(self):
        self.wb = None
        
    def create_excel_report(self, analysis_data: Dict) -> BytesIO:
        """Crea un reporte completo en Excel"""
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remover hoja por defecto
        
        # Crear hojas
        self._create_summary_sheet(analysis_data)
        self._create_liquidity_sheet(analysis_data)
        self._create_profitability_sheet(analysis_data)
        self._create_debt_sheet(analysis_data)
        self._create_rotation_sheet(analysis_data)
        self._create_bankruptcy_sheet(analysis_data)
        self._create_horizontal_analysis_sheet(analysis_data)
        self._create_vertical_analysis_sheet(analysis_data)
        
        # Guardar en memoria
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_summary_sheet(self, data: Dict):
        """Hoja de resumen ejecutivo"""
        ws = self.wb.create_sheet("Resumen Ejecutivo", 0)
        
        # Título
        ws['A1'] = "REPORTE DE ANÁLISIS FINANCIERO"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        years = data.get('available_years', [])
        indicators = data.get('indicators', {})
        
        row = 3
        ws[f'A{row}'] = "Período de Análisis"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{min(years)} - {max(years)}" if years else "N/A"
        
        row += 2
        ws[f'A{row}'] = "INDICADORES PRINCIPALES"
        ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        ws.merge_cells(f'A{row}:F{row}')
        
        # Indicadores clave del último año
        if years:
            last_year = str(max(years))
            row += 2
            
            # Liquidez
            self._add_indicator_row(ws, row, "Razón Corriente", 
                                   indicators.get('liquidez', {}).get('razon_corriente', {}).get(last_year, 0),
                                   "ratio")
            row += 1
            
            # Rentabilidad
            self._add_indicator_row(ws, row, "ROE", 
                                   indicators.get('rentabilidad', {}).get('roe', {}).get(last_year, 0),
                                   "percentage")
            row += 1
            self._add_indicator_row(ws, row, "ROA", 
                                   indicators.get('rentabilidad', {}).get('roa', {}).get(last_year, 0),
                                   "percentage")
            row += 1
            
            # Endeudamiento
            self._add_indicator_row(ws, row, "Endeudamiento Total", 
                                   indicators.get('endeudamiento', {}).get('endeudamiento_total', {}).get(last_year, 0),
                                   "percentage")
            row += 1
            
            # Z-Score
            z_score_data = indicators.get('quiebra', {}).get('z_score', {}).get(last_year, 0)
            z_score = z_score_data if isinstance(z_score_data, (int, float)) else 0
            self._add_indicator_row(ws, row, "Z-Score (Altman)", z_score, "number")
            
            clasificacion = indicators.get('quiebra', {}).get('clasificacion_z', {}).get(last_year, "N/A")
            row += 1
            ws[f'A{row}'] = "Estado Financiero"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = clasificacion
            
            # Color según clasificación
            if clasificacion == "Zona Segura":
                ws[f'B{row}'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            elif clasificacion == "Zona Gris":
                ws[f'B{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            else:
                ws[f'B{row}'].fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_liquidity_sheet(self, data: Dict):
        """Hoja de indicadores de liquidez"""
        ws = self.wb.create_sheet("Liquidez")
        self._create_indicator_sheet(ws, data, 'liquidez', 'INDICADORES DE LIQUIDEZ')
    
    def _create_profitability_sheet(self, data: Dict):
        """Hoja de indicadores de rentabilidad"""
        ws = self.wb.create_sheet("Rentabilidad")
        self._create_indicator_sheet(ws, data, 'rentabilidad', 'INDICADORES DE RENTABILIDAD')
    
    def _create_debt_sheet(self, data: Dict):
        """Hoja de indicadores de endeudamiento"""
        ws = self.wb.create_sheet("Endeudamiento")
        self._create_indicator_sheet(ws, data, 'endeudamiento', 'INDICADORES DE ENDEUDAMIENTO')
    
    def _create_rotation_sheet(self, data: Dict):
        """Hoja de indicadores de rotación"""
        ws = self.wb.create_sheet("Rotación")
        self._create_indicator_sheet(ws, data, 'rotacion', 'INDICADORES DE ROTACIÓN')
    
    def _create_bankruptcy_sheet(self, data: Dict):
        """Hoja de análisis de quiebra"""
        ws = self.wb.create_sheet("Análisis de Quiebra")
        self._create_indicator_sheet(ws, data, 'quiebra', 'INDICADORES DE QUIEBRA (Z-SCORE)')
    
    def _create_indicator_sheet(self, ws, data: Dict, indicator_type: str, title: str):
        """Plantilla genérica para hojas de indicadores"""
        # Título
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        years = data.get('available_years', [])
        indicators = data.get('indicators', {}).get(indicator_type, {})
        
        # Headers
        ws['A3'] = "Indicador"
        ws['A3'].font = Font(bold=True)
        ws['A3'].fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        
        for idx, year in enumerate(years):
            col = chr(66 + idx)  # B, C, D...
            ws[f'{col}3'] = str(year)
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            ws[f'{col}3'].alignment = Alignment(horizontal="center")
        
        # Datos
        row = 4
        for indicator_name, values in indicators.items():
            if isinstance(values, dict) and not indicator_name.startswith('clasificacion'):
                ws[f'A{row}'] = indicator_name.replace('_', ' ').title()
                
                for idx, year in enumerate(years):
                    col = chr(66 + idx)
                    value = values.get(str(year), 0)
                    
                    if isinstance(value, dict):
                        ws[f'{col}{row}'] = str(value)
                    else:
                        ws[f'{col}{row}'] = round(float(value), 2)
                        ws[f'{col}{row}'].number_format = '#,##0.00'
                
                row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        for idx in range(len(years)):
            col = chr(66 + idx)
            ws.column_dimensions[col].width = 15
        
        # Merge título
        ws.merge_cells(f'A1:{chr(65 + len(years))}1')
    
    def _create_horizontal_analysis_sheet(self, data: Dict):
        """Análisis horizontal (comparación entre períodos)"""
        ws = self.wb.create_sheet("Análisis Horizontal")
        
        ws['A1'] = "ANÁLISIS HORIZONTAL"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        ws.merge_cells('A1:E1')
        
        years = data.get('available_years', [])
        raw_data = data.get('raw_data', {})
        
        if len(years) < 2:
            ws['A3'] = "Se requieren al menos 2 períodos para análisis horizontal"
            return
        
        # Headers
        row = 3
        ws[f'A{row}'] = "Cuenta"
        ws[f'B{row}'] = f"Año {years[-2]}"
        ws[f'C{row}'] = f"Año {years[-1]}"
        ws[f'D{row}'] = "Variación Absoluta"
        ws[f'E{row}'] = "Variación %"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        
        # Datos
        row += 1
        for account, values in raw_data.items():
            year1 = values.get(str(years[-2]), 0)
            year2 = values.get(str(years[-1]), 0)
            
            if year1 != 0 or year2 != 0:
                ws[f'A{row}'] = account.replace('_', ' ').title()
                ws[f'B{row}'] = round(float(year1), 2)
                ws[f'C{row}'] = round(float(year2), 2)
                ws[f'D{row}'] = round(float(year2) - float(year1), 2)
                
                if year1 != 0:
                    var_percent = ((float(year2) / float(year1)) - 1) * 100
                    ws[f'E{row}'] = round(var_percent, 2)
                    ws[f'E{row}'].number_format = '0.00"%"'
                    
                    # Color según variación
                    if var_percent > 10:
                        ws[f'E{row}'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    elif var_percent < -10:
                        ws[f'E{row}'].fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                
                row += 1
        
        # Ajustar anchos
        for col, width in [('A', 30), ('B', 15), ('C', 15), ('D', 18), ('E', 15)]:
            ws.column_dimensions[col].width = width
    
    def _create_vertical_analysis_sheet(self, data: Dict):
        """Análisis vertical (estructura porcentual)"""
        ws = self.wb.create_sheet("Análisis Vertical")
        
        ws['A1'] = "ANÁLISIS VERTICAL"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        
        years = data.get('available_years', [])
        raw_data = data.get('raw_data', {})
        
        # Headers
        ws['A3'] = "Cuenta"
        for idx, year in enumerate(years):
            col = chr(66 + idx)
            ws[f'{col}3'] = f"% {year}"
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        
        ws['A3'].font = Font(bold=True)
        ws['A3'].fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        
        # Calcular porcentajes sobre activo total
        activo_total = raw_data.get('activo_total', {})
        
        row = 4
        for account, values in raw_data.items():
            if account != 'activo_total':
                ws[f'A{row}'] = account.replace('_', ' ').title()
                
                for idx, year in enumerate(years):
                    col = chr(66 + idx)
                    value = values.get(str(year), 0)
                    total = activo_total.get(str(year), 1)
                    
                    if total != 0:
                        percentage = (float(value) / float(total)) * 100
                        ws[f'{col}{row}'] = round(percentage, 2)
                        ws[f'{col}{row}'].number_format = '0.00"%"'
                
                row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        for idx in range(len(years)):
            col = chr(66 + idx)
            ws.column_dimensions[col].width = 15
        
        # Merge título
        ws.merge_cells(f'A1:{chr(65 + len(years))}1')
    
    def _add_indicator_row(self, ws, row, label, value, format_type="number"):
        """Agregar fila de indicador con formato"""
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        
        if format_type == "percentage":
            ws[f'B{row}'] = f"{round(float(value) * 100, 2)}%"
        elif format_type == "ratio":
            ws[f'B{row}'] = round(float(value), 2)
        else:
            ws[f'B{row}'] = round(float(value), 2)